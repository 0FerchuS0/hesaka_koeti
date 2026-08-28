import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _fmt_grad(value: float) -> str:
    return f"{value:+.2f}" if value else "PL"


def _fmt_adicion(value) -> str:
    return f"{value:+.2f}" if value else "-"


def generar_excel_reporte_graduaciones(datos, fecha_desde, fecha_hasta, tipo_lente: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Graduaciones"

    row = 1
    ws.cell(row=row, column=1, value="HESAKA").font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value="Graduaciones más demandadas (compra de stock de cristales)").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value=f"Periodo: {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')} — Tipo de lente: {tipo_lente}")
    row += 2

    headers_resumen = ["Unidades analizadas", "Excluidas (sin graduación)", "% Demanda terminado", "% Demanda laboratorio"]
    values_resumen = [
        datos.total_unidades_analizadas,
        datos.total_excluidas_sin_graduacion,
        datos.porcentaje_demanda_terminado,
        datos.porcentaje_demanda_laboratorio,
    ]
    for col, header in enumerate(headers_resumen, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        ws.cell(row=row + 1, column=col, value=values_resumen[col - 1])
    row += 4

    ws.cell(row=row, column=1, value="Ranking de graduaciones").font = Font(bold=True, size=11)
    row += 1
    headers = ["Esfera", "Cilindro", "Adición", "Recetadas", "Vendidas acá", "Receta externa", "Demanda total", "Tasa conversión", "% del total", "¿Terminado?"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111827")
    row += 1

    for fila in datos.filas:
        ws.cell(row=row, column=1, value=_fmt_grad(fila.esfera))
        ws.cell(row=row, column=2, value=_fmt_grad(fila.cilindro))
        ws.cell(row=row, column=3, value=_fmt_adicion(fila.adicion))
        ws.cell(row=row, column=4, value=fila.recetadas)
        ws.cell(row=row, column=5, value=fila.vendidas_aca)
        ws.cell(row=row, column=6, value=fila.receta_externa)
        ws.cell(row=row, column=7, value=fila.demanda_total)
        ws.cell(row=row, column=8, value=round(fila.tasa_conversion * 100, 1) if fila.tasa_conversion is not None else "-")
        ws.cell(row=row, column=9, value=round(fila.porcentaje_total, 1))
        ws.cell(row=row, column=10, value="Sí" if fila.es_terminado else "No")
        row += 1

    widths = {"A": 12, "B": 12, "C": 12, "D": 14, "E": 16, "F": 16, "G": 16, "H": 18, "I": 14, "J": 14}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
